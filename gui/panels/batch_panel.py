"""
Batch Panel - Batch Face Recognition
Feature #6: Process folder of images and export results to CSV
"""

import customtkinter as ctk
from tkinter import filedialog
import threading
from pathlib import Path
import pandas as pd
from datetime import datetime
import os

from config import Config
from utils.toast import ToastManager


class BatchPanel(ctk.CTkScrollableFrame):
    """Panel for batch face recognition from folder"""

    def __init__(self, parent, app):
        super().__init__(parent, fg_color="transparent")
        self.app = app
        self.toast = ToastManager(self)

        self.folder_path = None
        self.image_files = []
        self.results = []
        self.is_processing = False

        self.create_widgets()

    def create_widgets(self):
        """Create all UI widgets"""
        # Title
        title = ctk.CTkLabel(
            self,
            text="📁 Batch Face Recognition",
            font=ctk.CTkFont(size=Config.FONT_SIZE_TITLE, weight="bold")
        )
        title.pack(pady=(Config.PADDING_SMALL, Config.PADDING_TINY))

        subtitle = ctk.CTkLabel(
            self,
            text="Process multiple face images and export results to CSV/Excel",
            font=ctk.CTkFont(size=Config.FONT_SIZE_LABEL),
            text_color=Config.COLOR_INFO_TEXT
        )
        subtitle.pack(pady=(0, Config.PADDING_LARGE))

        # Configuration Frame
        config_frame = ctk.CTkFrame(self, corner_radius=15)
        config_frame.pack(fill="x", padx=Config.PADDING_LARGE, pady=Config.PADDING_SMALL)

        ctk.CTkLabel(
            config_frame,
            text="⚙️ Configuration",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SECTION, weight="bold")
        ).pack(pady=(Config.PADDING_MEDIUM, Config.PADDING_SMALL))

        # Folder selection
        folder_frame = ctk.CTkFrame(config_frame, fg_color="transparent")
        folder_frame.pack(fill="x", padx=Config.PADDING_LARGE, pady=Config.PADDING_SMALL)

        ctk.CTkLabel(
            folder_frame,
            text="Select Folder:",
            font=ctk.CTkFont(size=Config.FONT_SIZE_LABEL, weight="bold")
        ).pack(side="left", padx=(0, Config.PADDING_SMALL))

        self.folder_label = ctk.CTkLabel(
            folder_frame,
            text="No folder selected",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SMALL),
            text_color=Config.COLOR_INFO_TEXT
        )
        self.folder_label.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(
            folder_frame,
            text="📁 Browse",
            command=self.select_folder,
            width=120,
            height=35
        ).pack(side="right")

        # File count
        self.file_count_label = ctk.CTkLabel(
            config_frame,
            text="Images found: 0",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SMALL),
            text_color=Config.COLOR_INFO_TEXT
        )
        self.file_count_label.pack(pady=(Config.PADDING_TINY, Config.PADDING_MEDIUM))

        # Process button
        self.process_btn = ctk.CTkButton(
            self,
            text="▶️ Start Batch Processing",
            command=self.start_processing,
            width=300,
            height=60,
            font=ctk.CTkFont(size=Config.FONT_SIZE_SECTION, weight="bold"),
            fg_color=Config.COLOR_SUCCESS,
            hover_color=Config.COLOR_SUCCESS_HOVER,
            state="disabled"
        )
        self.process_btn.pack(pady=Config.PADDING_LARGE)

        # Progress Frame
        progress_frame = ctk.CTkFrame(self, corner_radius=15)
        progress_frame.pack(fill="x", padx=Config.PADDING_LARGE, pady=Config.PADDING_SMALL)

        ctk.CTkLabel(
            progress_frame,
            text="📊 Progress",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SECTION, weight="bold")
        ).pack(pady=(Config.PADDING_MEDIUM, Config.PADDING_SMALL))

        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(progress_frame, width=600)
        self.progress_bar.pack(padx=Config.PADDING_LARGE, pady=Config.PADDING_SMALL)
        self.progress_bar.set(0)

        # Progress label
        self.progress_label = ctk.CTkLabel(
            progress_frame,
            text="Ready to process",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SMALL)
        )
        self.progress_label.pack(pady=(0, Config.PADDING_MEDIUM))

        # Results Frame
        results_frame = ctk.CTkFrame(self, corner_radius=15)
        results_frame.pack(fill="both", expand=True, padx=Config.PADDING_LARGE,
                          pady=Config.PADDING_SMALL)

        ctk.CTkLabel(
            results_frame,
            text="📋 Results Summary",
            font=ctk.CTkFont(size=Config.FONT_SIZE_SECTION, weight="bold")
        ).pack(pady=(Config.PADDING_MEDIUM, Config.PADDING_SMALL))

        # Results text
        self.results_text = ctk.CTkTextbox(
            results_frame,
            height=200,
            font=ctk.CTkFont(size=Config.FONT_SIZE_SMALL)
        )
        self.results_text.pack(fill="both", expand=True,
                               padx=Config.PADDING_MEDIUM,
                               pady=(0, Config.PADDING_MEDIUM))
        self.results_text.insert("1.0", "No results yet. Select a folder and start processing.")
        self.results_text.configure(state="disabled")

        # Export buttons
        export_frame = ctk.CTkFrame(results_frame, fg_color="transparent")
        export_frame.pack(pady=(0, Config.PADDING_MEDIUM))

        self.export_csv_btn = ctk.CTkButton(
            export_frame,
            text="💾 Export to CSV",
            command=self.export_csv,
            width=180,
            height=Config.BUTTON_HEIGHT_MEDIUM,
            state="disabled"
        )
        self.export_csv_btn.pack(side="left", padx=Config.PADDING_TINY)

        self.export_excel_btn = ctk.CTkButton(
            export_frame,
            text="📊 Export to Excel",
            command=self.export_excel,
            width=180,
            height=Config.BUTTON_HEIGHT_MEDIUM,
            state="disabled",
            fg_color=Config.COLOR_SUCCESS,
            hover_color=Config.COLOR_SUCCESS_HOVER
        )
        self.export_excel_btn.pack(side="left", padx=Config.PADDING_TINY)

    def select_folder(self):
        """Select folder containing face images"""
        folder = filedialog.askdirectory(
            title="Select Folder with Face Images"
        )

        if folder:
            self.folder_path = folder
            self.scan_folder()

    def scan_folder(self):
        """Scan folder for image files"""
        if not self.folder_path:
            return

        # Find image files
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
        self.image_files = []

        for ext in image_extensions:
            self.image_files.extend(Path(self.folder_path).glob(f'*{ext}'))
            self.image_files.extend(Path(self.folder_path).glob(f'*{ext.upper()}'))

        # Update UI
        self.folder_label.configure(
            text=self.folder_path,
            text_color="white"
        )

        count = len(self.image_files)
        self.file_count_label.configure(
            text=f"Images found: {count}",
            text_color=Config.COLOR_SUCCESS_TEXT if count > 0 else Config.COLOR_ERROR_TEXT
        )

        if count > 0:
            self.process_btn.configure(state="normal")
            self.toast.show_success(
                f"Found {count} image(s) in selected folder",
                duration=Config.TOAST_DURATION_SUCCESS
            )
        else:
            self.toast.show_warning(
                "No image files found in selected folder!",
                duration=Config.TOAST_DURATION_WARNING
            )

    def start_processing(self):
        """Start batch processing"""
        if not self.image_files:
            self.toast.show_warning(
                "No images to process!",
                duration=Config.TOAST_DURATION_WARNING
            )
            return

        if self.is_processing:
            self.toast.show_info(
                "Already processing...",
                duration=Config.TOAST_DURATION_INFO
            )
            return

        self.is_processing = True
        self.process_btn.configure(state="disabled", text="⏸️ Processing...")
        self.export_csv_btn.configure(state="disabled")
        self.export_excel_btn.configure(state="disabled")
        self.results = []

        # Clear results
        self.results_text.configure(state="normal")
        self.results_text.delete("1.0", "end")
        self.results_text.insert("1.0", "Processing started...\n")
        self.results_text.configure(state="disabled")

        self.toast.show_info(
            "Starting batch processing...",
            duration=Config.TOAST_DURATION_INFO
        )

        # Process in thread
        def process_thread():
            total = len(self.image_files)
            recognized_count = 0
            unknown_count = 0
            error_count = 0

            for idx, img_path in enumerate(self.image_files):
                try:
                    # Update progress
                    progress = (idx + 1) / total
                    self.after(0, lambda p=progress, i=idx + 1, t=total:
                        self.update_progress(p, i, t))

                    # Recognize face
                    result = self.app.recognizer.recognize_face(str(img_path))

                    if result['recognized']:
                        person = result['person_info']
                        self.results.append({
                            'image_path': str(img_path),
                            'image_filename': Path(img_path).name,
                            'person_name': person['name'],
                            'person_id': person['face_id'],
                            'department': person.get('department', 'N/A'),
                            'role': person.get('role', 'N/A'),
                            'confidence': result['confidence'],
                            'status': 'Recognized'
                        })
                        recognized_count += 1
                    else:
                        self.results.append({
                            'image_path': str(img_path),
                            'image_filename': Path(img_path).name,
                            'person_name': 'Unknown',
                            'person_id': None,
                            'department': 'N/A',
                            'role': 'N/A',
                            'confidence': result['confidence'],
                            'status': 'Unknown'
                        })
                        unknown_count += 1

                except Exception as e:
                    self.results.append({
                        'image_path': str(img_path),
                        'image_filename': Path(img_path).name,
                        'person_name': 'Error',
                        'person_id': None,
                        'department': 'N/A',
                        'role': 'N/A',
                        'confidence': 0.0,
                        'status': f'Error: {str(e)}'
                    })
                    error_count += 1

            # Processing complete
            self.after(0, lambda: self.processing_complete(
                total, recognized_count, unknown_count, error_count
            ))

        thread = threading.Thread(target=process_thread, daemon=True)
        thread.start()

    def update_progress(self, progress, current, total):
        """Update progress bar and label"""
        self.progress_bar.set(progress)
        self.progress_label.configure(
            text=f"Processing: {current}/{total} images ({progress * 100:.1f}%)"
        )

    def processing_complete(self, total, recognized, unknown, errors):
        """Handle processing completion"""
        self.is_processing = False
        self.process_btn.configure(state="normal", text="▶️ Start Batch Processing")
        self.export_csv_btn.configure(state="normal")
        self.export_excel_btn.configure(state="normal")

        # Update results summary
        summary = f"""
╔══════════════════════════════════════════════════════╗
║           BATCH PROCESSING COMPLETE                  ║
╚══════════════════════════════════════════════════════╝

📊 Summary:
────────────────────────────────────────────────────────
Total Images Processed:  {total}
✅ Recognized:           {recognized} ({recognized / total * 100:.1f}%)
❌ Unknown:              {unknown} ({unknown / total * 100:.1f}%)
⚠️  Errors:               {errors} ({errors / total * 100:.1f}%)

🕒 Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
────────────────────────────────────────────────────────

Top Results:
"""

        # Show top 10 results
        for idx, result in enumerate(self.results[:10], 1):
            filename = result['image_filename']
            status = result['status']
            name = result['person_name']
            conf = result['confidence'] * 100

            summary += f"\n{idx}. {filename}"
            summary += f"\n   ➡️ {status}: {name} ({conf:.1f}%)"

        if len(self.results) > 10:
            summary += f"\n\n... and {len(self.results) - 10} more results"

        summary += "\n\n💾 Export results using buttons below."

        self.results_text.configure(state="normal")
        self.results_text.delete("1.0", "end")
        self.results_text.insert("1.0", summary)
        self.results_text.configure(state="disabled")

        # Show toast notification
        self.toast.show_success(
            f"Batch processing complete!\nProcessed: {total} | Recognized: {recognized} | Unknown: {unknown}",
            duration=Config.TOAST_DURATION_SUCCESS
        )

    def export_csv(self):
        """Export results to CSV"""
        if not self.results:
            self.toast.show_warning(
                "No results to export!",
                duration=Config.TOAST_DURATION_WARNING
            )
            return

        # Ask for save location
        filename = filedialog.asksaveasfilename(
            title="Save Results as CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"batch_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )

        if filename:
            try:
                # Create DataFrame with ordered columns
                df = pd.DataFrame(self.results)

                # Reorder columns for better readability
                column_order = [
                    'image_filename', 'person_name', 'person_id',
                    'department', 'role', 'confidence', 'status', 'image_path'
                ]
                df = df[column_order]

                # Export to CSV
                df.to_csv(filename, index=False)

                self.toast.show_success(
                    f"Results exported to CSV successfully!",
                    duration=Config.TOAST_DURATION_SUCCESS
                )

            except Exception as e:
                self.toast.show_error(
                    f"Failed to export CSV: {str(e)}",
                    duration=Config.TOAST_DURATION_ERROR
                )

    def export_excel(self):
        """Export results to Excel with colored formatting"""
        if not self.results:
            self.toast.show_warning(
                "No results to export!",
                duration=Config.TOAST_DURATION_WARNING
            )
            return

        # Ask for save location
        filename = filedialog.asksaveasfilename(
            title="Save Results as Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"batch_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if filename:
            try:
                # Create DataFrame with ordered columns
                df = pd.DataFrame(self.results)

                # Reorder columns for better readability
                column_order = [
                    'image_filename', 'person_name', 'person_id',
                    'department', 'role', 'confidence', 'status', 'image_path'
                ]
                df = df[column_order]

                # Create Excel writer
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Results')

                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Results']

                    # Add color coding based on confidence
                    from openpyxl.styles import PatternFill, Font, Alignment

                    # Define color fills
                    green_fill = PatternFill(
                        start_color='90EE90',
                        end_color='90EE90',
                        fill_type='solid'
                    )
                    yellow_fill = PatternFill(
                        start_color='FFFFE0',
                        end_color='FFFFE0',
                        fill_type='solid'
                    )
                    red_fill = PatternFill(
                        start_color='FFB6C1',
                        end_color='FFB6C1',
                        fill_type='solid'
                    )

                    # Header formatting
                    header_fill = PatternFill(
                        start_color='4472C4',
                        end_color='4472C4',
                        fill_type='solid'
                    )
                    header_font = Font(bold=True, color='FFFFFF')

                    # Apply header formatting
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Find confidence column index (1-based)
                    conf_col_idx = column_order.index('confidence') + 1
                    conf_col_letter = chr(64 + conf_col_idx)  # Convert to letter (A, B, C...)

                    # Apply colors to confidence column based on value
                    for row in range(2, len(df) + 2):
                        conf_val = df.iloc[row - 2]['confidence']
                        cell = worksheet[f'{conf_col_letter}{row}']

                        if conf_val >= 0.8:
                            cell.fill = green_fill
                        elif conf_val >= 0.6:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

                        # Format confidence as percentage
                        cell.number_format = '0.00%'

                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                self.toast.show_success(
                    "Results exported to Excel successfully!",
                    duration=Config.TOAST_DURATION_SUCCESS
                )

            except Exception as e:
                self.toast.show_error(
                    f"Failed to export Excel: {str(e)}",
                    duration=Config.TOAST_DURATION_ERROR
                )